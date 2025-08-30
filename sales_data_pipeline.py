# --- create_db.py ---
import sqlite3

def init_db():
    conn = sqlite3.connect('sales.db')
    cursor = conn.cursor()
    
    cursor.execute(''' 
    CREATE TABLE IF NOT EXISTS sales (
        id INTEGER PRIMARY KEY,
        date TEXT,
        product TEXT,
        quantity INTEGER,
        price REAL
    )
    '''
        )
    
    data = [
        ('2025-08-25', 'Widget A', 10, 9.99),
        ('2025-08-25', 'Widget B', 10, 19.99),
        ('2025-08-26', 'Widget A', 10, 9.99),
        ('2025-08-26', 'Widget C', 10, 29.99),
    ]
    cursor.executemany('INSERT INTO sales (date, product, quantity, price) VALUES(?,?,?,?)',data)
    
    conn.commit()
    conn.close()
    
    if __name__ == "__main__":
        init_db()
        
# --- fetch_data.py ---
import pandas as pd
import sqlite3

def get_sales_data():
    conn = sqlite3.connect('sales.db')
    df = pd.read_sql_query('SELECT * FROM sales', conn)
    conn.close()
    return df
    
# --- generate_report.py ---
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from fetch_data import get_sales_data

def generate_excel_report(filename='sales_report.xlsx'):
    df = get_sales_data()
    df['total'] = df['quantity'] * df['price']
    
    summary = df.groupby('product')['total'].sum().reset_index()
    
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Raw Data', index=False)
        summary.to_excel(writer, sheet_name='Summary', index=False)
        
        plt.figure()
        plt.bar(summary['product'], summary['total'])
        plt.title('Sales by Product')
        plt.xlabel('Product')
        plt.ylabel('Total Sales')
        
        chart_image = 'chart.png'
        plt.savefig(chart_image)
        
        wb = load_workbook(filename)
        ws = wb.create_sheet('Charts')
        from openpyxl.drawing.image import Image
        img = Image(chart_image)
        ws.add_image(img, 'A1')
        wb.save(filename)
        
# --- app.py ---
from flask import Flask, send_file
from generate_report import generate_excel_report

app = Flask(__name__)

@app.route('/report', methods=['GET'])
def get_report():
    filename = 'sales_report.xlsx'
    generate_excel_report(filename)
    return send_file(filename, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)